from unicodedata import name
from django.shortcuts import render

from apis.models import ApplicantExperienceModel, ApplicantInfoModel, ApplicantQualificationModel, CollegeInfoModel, EmployeeInfoModel, VacanciesInfoModel, VacancyApplicantMapping, RecruitmentCommitteeInfoModel

from django.contrib.auth.models import User
from django.http import HttpResponse, JsonResponse
from django.contrib.auth import login, authenticate, logout
from rest_framework.views import APIView
from rest_framework.response import Response
import openpyxl
from rest_framework.decorators import api_view
from django.core.mail import send_mail
from django.views.decorators.csrf import csrf_exempt
from datetime import *
# Create your views here.


@api_view(["GET"])
def dashboard_view(request, username):
    user = User.objects.get(username = username)
    college = CollegeInfoModel.objects.get(user = user)
    employee_count = EmployeeInfoModel.objects.filter(college = college).count()
    vacancies_count = VacanciesInfoModel.objects.filter(college = college).count()
    active_employee_count = EmployeeInfoModel.objects.filter(college = college, status = "Active").count()
    non_active_employee_count = EmployeeInfoModel.objects.filter(college = college, status = "Non Active").count()
    return Response({"employee_count": employee_count, "vacancies_count": vacancies_count, "active_employee_count": active_employee_count, "non_active_employee_count": non_active_employee_count}, status = 200)


# DOCUMENTATION DONE!
class ApplicantAPIView(APIView):
    # Method to fetch the details of an applicant using username.
    def get(self, request, username, format = None):
        user = User.objects.get(username = username)
        temp_result = {}
        temp_result["username"] = user.username
        temp_result["email"] = user.email
        if (ApplicantInfoModel.objects.filter(user = user).exists()):
            applicant = ApplicantInfoModel.objects.get(user = user)
            temp_result["profile_pic"] = applicant.profile_pic
            temp_result["description"] = applicant.description
            temp_result["full_name"] = applicant.full_name
            temp_result["DOB"] = applicant.DOB
            temp_result["gender"] = applicant.gender
            temp_result["address"] = applicant.address
            temp_result["state"] = applicant.state
            temp_result["pincode"] = applicant.pincode
            temp_result["category"] = applicant.category
            temp_result["marital_status"] = applicant.marital_status
            temp_result["phone_number"] = applicant.phone_number
            temp_result["total_experience"] = applicant.total_experience
            temp_result["skillset"] = applicant.skillset.names()
            temp_result["resume"] = applicant.resume
            return Response(temp_result, status = 200)
        else:
            return Response({"user" : temp_result, "mssg" : "Applicant profile not updated!"}, status = 403)

    # Posting the data of the applicant and user signup and login api.
    def post(self, request, format = None):
        if (request.data["purpose"] == "signup"):
            username = request.data["username"]
            if (User.objects.filter(username = username).exists()):
                return Response({"mssg" : "username already exists!"}, status = 409)
            email = request.data["email"]
            if (User.objects.filter(email = email).exists()):
                return Response({"mssg" : "email already exists!"}, status = 409)
            password = request.data["password"]
            confirm_password = request.data["confirm_password"]
            if (password == confirm_password):
                user = User.objects.create(username = username, email = email, password = password)
                user.set_password(user.password)
                user.save()
                login(request, user)
                return Response({"mssg": "user signed up successfully!"}, status = 200)
            else:
                return Response({"mssg": "passwords do not match!"}, status = 409)
        elif (request.data["purpose"] == "login"):
            if (authenticate(username = request.data["username"], password = request.data["password"])):
                return Response({"mssg": "user logedin successfully!"}, status = 200)
            else:
                return Response({"mssg": "login failed!"}, status = 404)
        elif (request.data["purpose"] == "fill details"):
            user = User.objects.get(username = request.data["username"])
            request.data["details"]["user"] = user
            temp = request.data["details"]["skillset"]
            del request.data["details"]["skillset"]
            applicant = ApplicantInfoModel.objects.create(**request.data["details"])
            for skill in temp:
                applicant.skillset.add(skill)
            # print (type(applicant.skillset))
            return Response({"mssg": "data updated successfully!"}, status = 202)

    # Updating the data of the applicant using username.
    def put(self, request, username, format = None):
        user = User.objects.get(username = username)
        if (ApplicantInfoModel.objects.filter(user = user).exists()):
            applicant = ApplicantInfoModel.objects.get(user = user)
            user.email = request.data.get("email")
            applicant.profile_pic = request.data.get("profile_pic")
            applicant.description = request.data.get("description")
            applicant.full_name = request.data.get("full_name")
            applicant.DOB = request.data.get("DOB")
            applicant.gender = request.data.get("gender")
            applicant.address = request.data.get("address")
            applicant.state = request.data.get("state")
            applicant.pincode = request.data.get("pincode")
            applicant.category = request.data.get("category")
            applicant.marital_status = request.data.get("marital_status")
            applicant.phone_number = request.data.get("phone_number")
            applicant.total_experience = request.data.get("total_experience")
            applicant.skillset.clear()
            for skill in request.data["skillset"]:
                applicant.skillset.add(skill)
            applicant.resume = request.data.get("resume")
            user.save()
            applicant.save()
            return Response({"mssg" : "user updated successfully"}, status = 204)
        else:
            return Response({"mssg : Personal Profile not updated!"}, status = 403)

    # Deleting user from the database.
    def delete(self, request, username, format = None):
        user = User.objects.get(username = username)
        user.delete()
        return Response({"mssg": "user delete successfully"}, status = 200)


class CollegeAPIView(APIView):
    def get(self, request, username, format = None):
        user = User.objects.get(username = username)
        if (CollegeInfoModel.objects.filter(user = user).exists()):
            college = CollegeInfoModel.objects.get(user = user)
            temp_result = {}
            temp_result["username"] = college.user.username
            temp_result["email"] = college.user.email
            temp_result["empid"] = college.empid
            temp_result["location"] = college.location
            temp_result["website"] = college.website
            temp_result["director_mail"] = college.director_mail
            temp_result["registrar_mail"] = college.registrar_mail
            temp_result["hod_mail"] = college.hod_mail
            return Response(temp_result, status = 200)
        else:
            return Response({"mssg" : "profile not updated!"}, status = 403)

    def post(self, request, format = None):
        if (request.data["purpose"] == "signup"):
            username = request.data["username"]
            if (User.objects.filter(username = username).exists()):
                return Response({"mssg" : "username already exists!"}, status = 409)
            email = request.data["email"]
            if (User.objects.filter(email = email).exists()):
                return Response({"mssg" : "email already exists!"}, status = 409)
            password = request.data["password"]
            confirm_password = request.data["confirm_password"]
            if (password == confirm_password):
                user = User.objects.create(username = username, email = email, password = password)
                user.set_password(user.password)
                user.save()
                login(request, user)
                return Response({"mssg": "user signed up successfully!"}, status = 200)
            else:
                return Response({"mssg": "passwords do not match!"}, status = 409)
        elif (request.data["purpose"] == "login"):
            if (authenticate(username = request.data["username"], password = request.data["password"])):
                return Response({"mssg": "user logedin successfully!"}, status = 200)
            else:
                return Response({"mssg": "login failed!"}, status = 404)
        elif (request.data["purpose"] == "fill details"):
            user = User.objects.get(username = request.data["username"])
            request.data["details"]["user"] = user
            CollegeInfoModel.objects.create(**request.data["details"])
            return Response({"mssg": "data updated successfully!"}, status = 202)

    def put(self, request, username, format = None):
        user = User.objects.get(username = username)
        college = CollegeInfoModel.objects.get(user = user)
        user.email = request.data.get("email")
        college.empid = request.data.get("empid")
        college.location = request.data.get("location")
        college.website = request.data.get("website")
        college.director_mail = request.data.get("director_mail")
        college.registrar_mail = request.data.get("registrar_mail")
        college.hod_mail = request.data.get("hod_mail")
        user.save()
        college.save()
        return Response({"mssg" : "user updated successfully"}, status = 204)

    def delete(self, request, username, format = None):
        user = User.objects.get(username = username)
        college = CollegeInfoModel.objects.get(user = user)
        user.delete()
        college.delete()
        return Response({"mssg": "user delete successfully"}, status = 200)


# DOCUMENTATION DONE!
# Model for performing the CRUD operations on the user qualification.
class QualificationAPIView(APIView):

    # Method to get the qualifications of a particular applicant using username.
    def get(self, request, username, format = None):
        applicant = User.objects.get(username = username)
        qualifications = ApplicantQualificationModel.objects.filter(applicant = applicant)
        result = []
        for qualification in qualifications:
            temp_result = {}
            temp_result["qualification_title"] = qualification.qualification_title
            temp_result["institute"] = qualification.institute
            temp_result["passing_year"] = qualification.passing_year
            temp_result["marks"] = qualification.marks
            result.append(temp_result)
        return Response({"data" : result}, status = 200)

    # Posting the qualifications of an applicant using the username of the applicant.
    def post(self, request, username, format = None):
        user = User.objects.get(username = username)
        request.data["applicant"] = user
        ApplicantQualificationModel.objects.create(**request.data)
        return Response({"mssg" : "qualification added successfully"}, status = 202)

    # Method to update the qualifications of a user using the username and the qualification title.
    def put(self, request, username, format = None):
        applicant = User.objects.get(username = username)
        qualification = ApplicantQualificationModel.objects.get(applicant = applicant, qualification_title = request.data["title"])
        qualification.qualification_title = request.data["qualification_title"]
        qualification.institute = request.data["institute"]
        qualification.passing_year = request.data["passing_year"]
        qualification.marks = request.data["marks"]
        qualification.save()
        return Response({"mssg" : "qualification updated successfully"}, status = 204)

    # deleting the qualification of the user using username and the qualification title.
    def delete(self, request, username, format = None):
        applicant = User.objects.get(username = username)
        qualification = ApplicantQualificationModel.objects.get(applicant = applicant, qualification_title = request.data["qualification_title"])
        qualification.delete()
        return Response({"mssg": "qualification deleted successfully"}, status = 200)


# DOCUMENTATION DONE!
# Method to get the employee data using the college name and the employee id. We need to mention get in the square brackets else nothing will work.
@api_view(["GET"])
def get_employee_by_id(request, college_name, id):
    if (request.method == "GET"):
        user = User.objects.get(username = college_name)
        college = CollegeInfoModel.objects.get(user = user)
        employee = EmployeeInfoModel.objects.get(college = college, id = id)
        temp_result = employee.__dict__
        del temp_result["_state"]
        return Response({"employee" : temp_result}, status = 200)


@api_view(["GET"])
def get_employee_by_empid(request, college_name, empid):
    if (request.method == "GET"):
        print (college_name, empid)
        user = User.objects.get(username = college_name)
        college = CollegeInfoModel.objects.get(user = user)
        if (EmployeeInfoModel.objects.filter(college = college, empid = empid).exists()):
            employee = EmployeeInfoModel.objects.get(college = college, empid = empid)
            temp_result = employee.__dict__
            del temp_result["_state"]
            return Response({"employee" : temp_result}, status = 200)
        else:
            return Response({"mssg" : "employee does not exists!"}, status = 404)


# DOCUMENTATION DONE!
# this api is used to change the status of the employee from active to notice period and mail all the required faculties that recruitment process has been initiated.
# @api_view(["POST"])
# def Change_employee_status(request, college_name, id):
#     if (request.method == "POST"):
#         user = User.objects.get(username = college_name)
#         college = CollegeInfoModel.objects.get(user = user)
#         employee = EmployeeInfoModel.objects.get(college = college, id = id)
#         employee.status = request.data["status"]
#         employee.save()
#         message_name = "Initiate recruitment process"
#         message_email = "rapidrecruits1.0@gmail.com"
#         message = "Dear all, Mr./Mrs. {} is about to leave/retire from their position please initiate recruitment process".format(employee.name)
#         send_mail(
#             message_name,#subject
#             message,#message
#             message_email,#from email
#             [college.director_mail, college.registrar_mail, college.hod_mail, employee.email, college.user.email],#to email
#         )
#         return Response({"mssg": "status changed successfully!"}, status = 204)


# DOCUMENTATION DONE!
class EmployeeAPIView(APIView):

    # Method to fetch the data of all the employees of the particular college using college name.
    def get(self, request, college_name, format = None):
        user = User.objects.get(username = college_name)
        college = CollegeInfoModel.objects.get(user = user)
        employees = EmployeeInfoModel.objects.filter(college = college)
        result = []
        for employee in employees:
            temp_result = {}
            temp = employee.__dict__
            print (temp)
            for key in temp:
                # This state is the reference object to the college.
                if (key == "_state"):
                    continue
                temp_result[key] = temp[key]
            result.append(temp_result)
            dob_year = (datetime.strptime(employee.DOB,"%d/%m/%Y")).year
            dob_month = (datetime.strptime(employee.DOB,"%d/%m/%Y")).month
            dob_date = (datetime.strptime(employee.DOB,"%d/%m/%Y")).day
            # print(dob_year,dob_date,dob_month,date.today().year,date.today().day,date.today().month)
            if dob_month >=1 and dob_month<=3 and date.today().day == dob_date and date.today().year == dob_year + 59 and date.today().month == dob_month + 9 :
                message_name = "Initiate recruitment process"
                message_email = "rapidrecruits1.0@gmail.com"
                message = "Dear all, Mr./Mrs. {} is about to leave/retire from their position please initiate recruitment process".format(employee.name)
                send_mail(
                    message_name,#subject
                    message,#message
                    message_email,#from email
                    [college.director_mail, college.registrar_mail, college.hod_mail, employee.email, college.user.email],#to email
                )
            elif dob_month >=4 and dob_month<=12 and date.today().day == dob_date and date.today().year == dob_year + 60 and date.today().month == dob_month - 3:
                message_name = "Initiate recruitment process"
                message_email = "rapidrecruits1.0@gmail.com"
                message = "Dear all, Mr./Mrs. {} is about to leave/retire from their position please initiate recruitment process".format(employee.name)
                send_mail(
                    message_name,#subject
                    message,#message
                    message_email,#from email
                    [college.director_mail, college.registrar_mail, college.hod_mail, employee.email, college.user.email],#to email
                )
        result = sorted(result, key = lambda x : x["name"].lower())
        return Response({"employees" : result}, status = 200)

    # Method to post the data of the employees of a college using excel sheet or by using manual method using college name.
    def post(self, request, college_name, format = None):
        user = User.objects.get(username = college_name)
        college = CollegeInfoModel.objects.get(user = user)
        if (request.data["method"] == "excel file"):
            wb_obj = openpyxl.load_workbook(request.FILES["details"])
            sheet_obj = wb_obj.active
            row = sheet_obj.max_row
            column = sheet_obj.max_column
            count = 0
            print (row, column, sheet_obj.cell(row = 2, column = 1).value)
            keys = ["college", "empid", "name", "DOB", "gender", "category", "status", "designation", "department", "email", "phone_number"]
            for i in range(2, row + 1):
                values = [college]
                for j in range(1, column + 1):
                    values.append(sheet_obj.cell(row = i, column = j).value)
                comb_list = zip(keys, values)
                data = dict(comb_list)
                if (data["name"] == None):
                    break
                EmployeeInfoModel.objects.create(**data)
                count += 1
            return Response({"mssg": "{} number of records created".format(count)}, status = 201)
        elif (request.data["method"] == "manual"):
            request.data["details"]["college"] = college
            EmployeeInfoModel.objects.create(**request.data["details"])
            return Response({"mssg": "employee added successfully!"}, status = 201)

    # Method to update the data of the employee using college name a.k.a. username and the employee id as there is no other unique parameter to be used.
    def put(self, request, college_name, format = None):
        user = User.objects.get(username = college_name)
        college = CollegeInfoModel.objects.get(user = user)
        employee = EmployeeInfoModel.objects.get(college = college, id = request.data["id"])
        flag = 0
        if (employee.status == "Non Active"):
            flag = 1
        employee.name = request.data["name"]
        employee.DOB = request.data["DOB"]
        employee.gender = request.data["gender"]
        employee.category = request.data["category"]
        employee.status = request.data["status"]
        employee.email = request.data["email"]
        employee.phone_number = request.data["phone_number"]
        employee.designation = request.data["designation"]
        employee.empid = request.data["empid"]
        employee.department = request.data["department"]
        employee.save()
        if (request.data["status"] == "Non Active" and flag == 0):
            message_name = "Initiate recruitment process"
            message_email = "rapidrecruits1.0@gmail.com"
            message = "Dear all, Mr./Mrs. {} is about to leave/retire from their position please initiate recruitment process".format(employee.name)
            send_mail(
                message_name,#subject
                message,#message
                message_email,#from email
                [college.director_mail, college.registrar_mail, college.hod_mail, employee.email, college.user.email],#to email
            )
        return Response({"mssg": "employee details updated successfully"}, status = 204)

    # Method to delete the record of an employee using college name and id of the employee.
    def delete(self, request, college_name, format = None):
        user = User.objects.get(username = college_name)
        college = CollegeInfoModel.objects.get(user = user)
        employee = EmployeeInfoModel.objects.get(college = college, id = request.data["id"])
        employee.delete()
        return Response({"mssg": "employee deleted successfully!"}, status = 200)

# DOCUMENTATION DONE!
# Method to get the vacancies for a particular applicant where the applicant has applied.
@api_view(["GET"])
def get_vacancies_for_applicant(request, username):
    user = User.objects.get(username = username)
    mappings = VacancyApplicantMapping.objects.filter(applicant = user).order_by('date_of_application')
    vacancies = []
    for mapping in mappings:
        vacancies.append(mapping.vacancy)
    result = []
    for vacancy in vacancies:
        temp_result = {}
        temp = vacancy.__dict__
        # print (temp)
        for key in temp:
            # This state is the reference object to the college.
            if (key == "_state"):
                continue
            temp_result[key] = temp[key]
        temp_result["skills"] = vacancy.skills.names()
        temp_result["college_name"] = vacancy.college.user.username
        temp_result["location"] = vacancy.college.location
        temp_result["website"] = vacancy.college.website
        result.append(temp_result)
    return Response({"vacancies" : result}, status = 200)

# DOCUMENTATION DONE!
# Method to get the applicants who have applied for a particular vacancy using vacancy id.
@api_view(["GET"])
def get_applicants_for_vacancy(request, id):
    mappings = VacancyApplicantMapping.objects.filter(vacancy = id).order_by('date_of_application')
    applicants = []
    status = {}
    for mapping in mappings:
        applicants.append(mapping.applicant)
        status[mapping.applicant] = mapping.status
    result = []
    for applicant in applicants:
        applicant_info_dict = {}
        personal_details_dict = {}
        personal_details_dict["username"] = applicant.username
        personal_details_dict["email"] = applicant.email
        # Getting the personal information of the applicant.
        personal_info_obj = ApplicantInfoModel.objects.get(user = applicant)
        temp = personal_info_obj.__dict__
        for key in temp:
            if (key == "_state"):
                continue
            personal_details_dict[key] = temp[key]
        personal_details_dict["skillset"] = personal_info_obj.skillset.names()
        applicant_info_dict["personal details"] = personal_details_dict
        # Getting the qualifications of the applicant.
        qualifications = ApplicantQualificationModel.objects.filter(applicant = applicant)
        all_qualifications = []
        for qualification in qualifications:
            temp_qualification = {}
            temp_qualification["qualification_title"] = qualification.qualification_title
            temp_qualification["institute"] = qualification.institute
            temp_qualification["passing_year"] = qualification.passing_year
            temp_qualification["marks"] = qualification.marks
            all_qualifications.append(temp_qualification)
        applicant_info_dict["qualification details"] = all_qualifications
        # Getting the experience of the applicant.
        experiences = ApplicantExperienceModel.objects.filter(applicant = applicant)
        all_experiences = []
        for experience in experiences:
            temp_experience = {}
            temp_experience["designation"] = experience.designation
            temp_experience["from_date"] = experience.from_date
            temp_experience["to_date"] = experience.to_date
            temp_experience["institute"] = experience.institute
            temp_experience["details"] = experience.details
            all_experiences.append(temp_experience)
        applicant_info_dict["experience details"] = all_experiences
        # Getting the status of employee for the vacancy.
        applicant_info_dict["status"] = status[applicant]
        result.append(applicant_info_dict)
    return Response({"applicants" : result}, status = 200)


# Method to get all the vacancies which require similar skills which the user has removed already applied vacancies..
@api_view(["GET"])
def search_matching_vacancies(request, username):
    user = User.objects.get(username = username)
    applicant = ApplicantInfoModel.objects.get(user = user)
    # Here, we are getting the id's of the vacancies which the applicant has already applied for.
    applied_vacancies_mapping = VacancyApplicantMapping.objects.filter(applicant = user)
    applied_vacancies_id = []
    for mapping in applied_vacancies_mapping:
        applied_vacancies_id.append(mapping.vacancy.id)
    vacancies = VacanciesInfoModel.objects.filter(skills__name__in = applicant.skillset.names()).distinct()
    # Now, we will exclude those vacancies from the resultant query set which the applicant has already applied for.
    vacancies = vacancies.exclude(id__in = applied_vacancies_id)
    result = []
    for vacancy in vacancies:
        temp_result = {}
        temp = vacancy.__dict__
        # print (temp)
        for key in temp:
            # This state is the reference object to the college.
            if (key == "_state"):
                continue
            temp_result[key] = temp[key]
        temp_result["skills"] = vacancy.skills.names()
        temp_result["college_name"] = vacancy.college.user.username
        temp_result["location"] = vacancy.college.location
        temp_result["website"] = vacancy.college.website
        result.append(temp_result)
    return Response({"vacancies" : result}, status = 200)


# Method to search applicants based on the skills required by the vacancies.
@api_view(["GET"])
def search_matching_applicants(request, id):
    vacancy = VacanciesInfoModel.objects.get(id = id)
    applicants_obj = ApplicantInfoModel.objects.filter(skillset__name__in = vacancy.skills.names()).distinct()
    # Get the applicants who have already applied for the vacancy.
    mapping_queryset = VacancyApplicantMapping.objects.filter(vacancy = vacancy)
    applied_users = []
    for mapping in mapping_queryset:
        applied_users.append(mapping.applicant)
    # Removing the applicants who have already applied for the vacancy.
    applicants_obj = applicants_obj.exclude(user__in = applied_users)

    applicants = []
    for applicant_obj in applicants_obj:
        applicants.append(User.objects.get(username = applicant_obj.user.username))
    result = []
    for applicant in applicants:
        applicant_info_dict = {}
        personal_details_dict = {}
        personal_details_dict["username"] = applicant.username
        personal_details_dict["email"] = applicant.email
        # Getting the personal information of the applicant.
        personal_info_obj = ApplicantInfoModel.objects.get(user = applicant)
        temp = personal_info_obj.__dict__
        for key in temp:
            if (key == "_state"):
                continue
            personal_details_dict[key] = temp[key]
        applicant_info_dict["personal details"] = personal_details_dict
        personal_details_dict["skillset"] = personal_info_obj.skillset.names()
        # Getting the qualifications of the applicant.
        qualifications = ApplicantQualificationModel.objects.filter(applicant = applicant)
        all_qualifications = []
        for qualification in qualifications:
            temp_qualification = {}
            temp_qualification["qualification_title"] = qualification.qualification_title
            temp_qualification["institute"] = qualification.institute
            temp_qualification["passing_year"] = qualification.passing_year
            temp_qualification["marks"] = qualification.marks
            all_qualifications.append(temp_qualification)
        applicant_info_dict["qualification details"] = all_qualifications
        # Getting the experience of the applicant.
        experiences = ApplicantExperienceModel.objects.filter(applicant = applicant)
        all_experiences = []
        for experience in experiences:
            temp_experience = {}
            temp_experience["designation"] = experience.designation
            temp_experience["from_date"] = experience.from_date
            temp_experience["to_date"] = experience.to_date
            temp_experience["institute"] = experience.institute
            temp_experience["details"] = experience.details
            all_experiences.append(temp_experience)
        applicant_info_dict["experience details"] = all_experiences
        result.append(applicant_info_dict)
    return Response({"applicants" : result}, status = 200)


@api_view(["GET"])
def get_vacancy_by_id(request, id):
    vacancy = VacanciesInfoModel.objects.get(id = id)
    temp_result = {}
    temp = vacancy.__dict__
    # print (temp)
    for key in temp:
        # This state is the reference object to the college.
        if (key == "_state"):
            continue
        temp_result[key] = temp[key]
    temp_result["skills"] = vacancy.skills.names()
    temp_result["college_name"] = vacancy.college.user.username
    temp_result["location"] = vacancy.college.location
    temp_result["website"] = vacancy.college.website
    flag = False
    if (RecruitmentCommitteeInfoModel.objects.filter(vacancy = vacancy).exists()):
        flag = True
    temp_result ["recruitment_committee"] = flag
    return Response({"vacancy" : temp_result}, status = 200)

# removed already applied vacancies.
@api_view(["GET"])
def get_all_vacancies_for_applicant(request, username):
        vacancies = VacanciesInfoModel.objects.all()
        user = User.objects.get(username = username)
        # Here, we are getting the id's of the vacancies which the applicant has already applied for.
        applied_vacancies_mapping = VacancyApplicantMapping.objects.filter(applicant = user)
        applied_vacancies_id = []
        for mapping in applied_vacancies_mapping:
            applied_vacancies_id.append(mapping.vacancy.id)
        # Now, we will exclude those vacancies from the resultant query set which the applicant has already applied for.
        vacancies = vacancies.exclude(id__in = applied_vacancies_id)
        result = []
        for vacancy in vacancies:
            temp_result = {}
            temp = vacancy.__dict__
            print (temp)
            for key in temp:
                print (key)
                # This state is the reference object to the college.
                if (key == "_state"):
                    continue
                temp_result[key] = temp[key]
            temp_result["skills"] = vacancy.skills.names()
            temp_result["college_name"] = vacancy.college.user.username
            temp_result["location"] = vacancy.college.location
            temp_result["website"] = vacancy.college.website
            flag = False
            if (RecruitmentCommitteeInfoModel.objects.filter(vacancy = vacancy).exists()):
                flag = True
            temp_result ["recruitment_committee"] = flag
            result.append(temp_result)
        return Response({"vacancies" : result}, status = 200)



# DOCUMENTATION DONE!
@api_view(["POST"])
def apply_for_vacancy(request, username):
    request.data["applicant"] = User.objects.get(username = username)
    request.data["vacancy"] = VacanciesInfoModel.objects.get(id = request.data["id"])
    request.data["status"] = "under review"
    del request.data["id"]
    request.data["date_of_application"] = datetime.now()
    VacancyApplicantMapping.objects.create(**request.data)
    return Response({"mssg" : "Applied for the vacancy successfully!"}, status = 200)


# Using this function we can change acncstatus of the applicant for a particular vacacncy.
@api_view(["PUT"])
def change_status_of_applicant(request, id, username):
    vacancy = VacanciesInfoModel.objects.get(id = id)
    applicant = User.objects.get(username = username)
    mapping = VacancyApplicantMapping.objects.get(vacancy = vacancy, applicant = applicant)
    mapping.status = request.data["status"]
    mapping.save()
    if(mapping.status == "rejected"):
        message_name = "Sorry"
        message_email = "rapidrecruits1.0@gmail.com"
        message = "Rejected"
        send_mail(
            message_name,#subject
            message,#message
            message_email,#from email
            [applicant.email]#to email
        )
    elif(mapping.status == "hired"):
        message_name = "Congratulations"
        message_email = "rapidrecruits1.0@gmail.com"
        message = "Selected"
        send_mail(
            message_name,#subject
            message,#message
            message_email,#from email
            [applicant.email]#to email
        )
    elif(mapping.status == "meet scheduled"):
        message_name = "Meeting Scheduled"
        message_email = "rapidrecruits1.0@gmail.com"
        message = "Congratulations ,  you are shortlisted. Meeting details is shared below "
        send_mail(
            message_name,#subject
            message,#message
            message_email,#from email
            [applicant.email]#to email
        )
    return Response({"mssg" : "Status updated successfully"}, status = 200)


# Using this API we can approach the applicant we are interested in.
@api_view(["POST"])
def approach_applicant(request, username):
    user = User.objects.get(username = username)
    emailid = user.email
    # print (user.email)
    link = request.data["link"]
    vacancy_id = request.data["id"]
    message_name = "Member of recruitement Committee"
    message_email = "rapidrecruits1.0@gmail.com"
    message = "Your profile looks suitable for this vacancy. If interested please apply at the link {} , Vacancy Id {}".format(link, vacancy_id)
    send_mail(
        message_name,#subject
        message,#message
        message_email,#from email
        [emailid]#to email
    )
    return Response({"mssg": "mail sent successfully"}, status = 200)

# DOCUMENTATION DONE!
class VacanciesAPIView(APIView):

    # Method to get all the vacancies posted by a particular college using college name or to get all the vacancies present in the system if college not given.
    def get(self, request, college_name = None, format = None):
        vacancies = VacanciesInfoModel.objects.all()
        if (college_name):
            user = User.objects.get(username = college_name)
            college = CollegeInfoModel.objects.get(user = user)
            vacancies = vacancies.filter(college = college)
        result = []
        for vacancy in vacancies:
            temp_result = {}
            temp = vacancy.__dict__
            print (temp)
            for key in temp:
                print (key)
                # This state is the reference object to the college.
                if (key == "_state"):
                    continue
                temp_result[key] = temp[key]
            temp_result["skills"] = vacancy.skills.names()
            temp_result["college_name"] = vacancy.college.user.username
            temp_result["location"] = vacancy.college.location
            temp_result["website"] = vacancy.college.website
            flag = False
            if (RecruitmentCommitteeInfoModel.objects.filter(vacancy = vacancy).exists()):
                flag = True
            temp_result ["recruitment_committee"] = flag
            result.append(temp_result)
        return Response({"vacancies" : result}, status = 200)

    # Method to post a new vacancy using college name or user name of the college.
    def post(self, request, college_name, format = None):
        user = User.objects.get(username = college_name)
        college = CollegeInfoModel.objects.get(user = user)
        employees = EmployeeInfoModel.objects.filter(college = college)
        request.data["college"] = college
        temp = request.data["skills"]
        del request.data["skills"]
        vacancy = VacanciesInfoModel.objects.create(**request.data)
        for skill in temp:
            vacancy.skills.add(skill)
        for employee in employees:
            message_name = "Vacancy Created"
            message_email = "rapidrecruits1.0@gmail.com"
            message = "Dear all, Vacancy for {} has been created".format(vacancy.title)
            send_mail(
                message_name,#subject
                message,#message
                message_email,#from email
                [employee.email],#to email   
            )
        return Response({"mssg": "Vacancy posted successfully!"}, status = 201)

    # Method to update the details of a particular vacancy using college name and id of the vacancy.
    def put(self, request, college_name, format = None):
        user = User.objects.get(username = college_name)
        college = CollegeInfoModel.objects.get(user = user)
        vacancy = VacanciesInfoModel.objects.get(college = college, id = request.data["id"])
        vacancy.title = request.data["title"]
        vacancy.type = request.data["type"]
        vacancy.experience = request.data["experience"]
        vacancy.date_of_posting = request.data["date_of_posting"]
        vacancy.state = request.data["state"]
        vacancy.description = request.data["description"]
        vacancy.responsibilities = request.data["responsibilities"]
        vacancy.qualifications = request.data["qualifications"]
        vacancy.skills.clear()
        for skill in request.data["skills"]:
            vacancy.skills.add(skill)
        vacancy.compensation = request.data["compensation"]
        vacancy.save()
        return Response({"mssg": "vacancy details updated successfully!"}, status = 204)

    # Method to delete a particular vacancy using college name and id.
    def delete(self, request, college_name, format = None):
        user = User.objects.get(username = college_name)
        college = CollegeInfoModel.objects.get(user = user)
        vacancy = VacanciesInfoModel.objects.get(college = college, id = request.data["id"])
        vacancy.delete()
        return Response({"mssg": "Vacancy deleted successfully!"}, status = 200)


class ExperienceAPIView(APIView):

    def get(self, request, username, format = None):
        applicant = User.objects.get(username = username)
        experiences = ApplicantExperienceModel.objects.filter(applicant = applicant)
        result = []
        for experience in experiences:
            temp_result = {}
            temp_result["designation"] = experience.designation
            temp_result["from_date"] = experience.from_date
            temp_result["to_date"] = experience.to_date
            temp_result["institute"] = experience.institute
            temp_result["details"] = experience.details
            result.append(temp_result)
        return Response({"data" : result}, status = 200)

    def post(self, request, username, format = None):
        user = User.objects.get(username = username)
        request.data["applicant"] = user
        ApplicantExperienceModel.objects.create(**request.data)
        return Response({"mssg" : "experience added successfully"}, status = 202)


    def put(self, request, username, format = None):
        applicant = User.objects.get(username = username)
        experience = ApplicantExperienceModel.objects.get(applicant = applicant, institute = request.data["institute"])
        experience.designation = request.data["designation"]
        experience.from_date = request.data["from_date"]
        experience.to_date = request.data["to_date"]
        experience.institute = request.data["institute"]
        experience.details = request.data["details"]
        experience.save()
        return Response({"mssg" : "experience updated successfully"}, status = 204)


    def delete(self, request, username, format = None):
        applicant = User.objects.get(username = username)
        experience = ApplicantExperienceModel.objects.get(applicant = applicant, institute = request.data["institute"])
        experience.delete()
        return Response({"mssg": "experience deleted successfully"}, status = 200)


# RESTRICTED SNIPPETS!

        # if (pk == None):
        #     applicants = ApplicantInfoModel.objects.all()
        #     result = []
        #     for applicant in applicants:
        #         temp_result = {}
        #         temp_result["id"] =
        #         temp_result["username"] = applicant.user.username
        #         temp_result["email"] = applicant.user.email
        #         temp_result["description"] = applicant.description
        #         temp_result["full_name"] = applicant.full_name
        #         temp_result["DOB"] = applicant.DOB
        #         temp_result["gender"] = applicant.gender
        #         temp_result["address"] = applicant.address
        #         temp_result["state"] = applicant.state
        #         temp_result["pincode"] = applicant.pincode
        #         temp_result["category"] = applicant.category
        #         temp_result["marital_status"] = applicant.marital_status
        #         temp_result["phone_number"] = applicant.phone_number
        #         temp_result["total_experience"] = applicant.total_experience
        #         temp_result["skillset"] = applicant.skillset
        #         result.append(temp_result)
        #     return Response(result, status = 200)

        # else:

#Send Email

class RecruitmentCommitteeAPIView(APIView):
    def get(self, request, college_name, id, format = None):
        vacancy=VacanciesInfoModel.objects.get(id=id)
        committee=RecruitmentCommitteeInfoModel.objects.get(vacancy=vacancy)
        first_user = committee.first_user
        second_user = committee.second_user
        third_user = committee.third_user
        fourth_user = committee.fourth_user
        fifth_user = committee.fifth_user
        # id,empid,name,designation,department
        committee_members=[]
        committee_members.append(first_user)
        committee_members.append(second_user)
        committee_members.append(third_user)
        committee_members.append(fourth_user)
        committee_members.append(fifth_user)
        result = []
        for member in committee_members:
            temp_result = {}
            temp_result["id"]=member.id
            temp_result["empid"]=member.empid
            temp_result["name"]=member.name
            temp_result["email"]=member.email   
            temp_result["phone_number"]=member.phone_number      
            temp_result["designation"]=member.designation
            temp_result["department"]=member.department
            result.append(temp_result)
        return Response({"data" : result}, status = 200)


    def post(self,request,college_name,id,format = None):
        first_user=EmployeeInfoModel.objects.get(id=request.data["first"])
        second_user=EmployeeInfoModel.objects.get(id=request.data["second"])
        third_user=EmployeeInfoModel.objects.get(id=request.data["third"])
        fourth_user=EmployeeInfoModel.objects.get(id=request.data["forth"])
        fifth_user=EmployeeInfoModel.objects.get(id=request.data["fifth"])
        vacancy=VacanciesInfoModel.objects.get(id=id)
        result={"first_user":first_user,"second_user":second_user,"third_user":third_user,"fourth_user":fourth_user,"fifth_user":fifth_user,"vacancy":vacancy}
        RecruitmentCommitteeInfoModel.objects.create(**result)
        message_name = "Member of recruitement Committee"
        message_email = "rapidrecruits1.0@gmail.com"
        message = "Dear all, You have been added as the member of Recruitment Committee for the Vacancy {} , Vacancy Id {}".format(vacancy.title,vacancy.id)
        send_mail(
            message_name,#subject
            message,#message
            message_email,#from email
            [first_user.email,second_user.email,third_user.email,fourth_user.email,fifth_user.email,]#to email
        )
        return Response({"mssg" : "committee added successfully"}, status = 202)


    def put(self, request, college_name, id, format = None):
        first_user = EmployeeInfoModel.objects.get(id = request.data["first"])
        second_user = EmployeeInfoModel.objects.get(id = request.data["second"])
        third_user = EmployeeInfoModel.objects.get(id = request.data["third"])
        fourth_user = EmployeeInfoModel.objects.get(id = request.data["forth"])
        fifth_user = EmployeeInfoModel.objects.get(id = request.data["fifth"])
        vacancy = VacanciesInfoModel.objects.get(id = id)
        recruitment_committee = RecruitmentCommitteeInfoModel.objects.get(vacancy = vacancy)
        recruitment_committee.delete()
        result={"first_user" : first_user, "second_user" : second_user, "third_user" : third_user, "fourth_user" : fourth_user, "fifth_user" : fifth_user, "vacancy" : vacancy}
        RecruitmentCommitteeInfoModel.objects.create(**result)
        message_name = "Member of recruitement Committee"
        message_email = "rapidrecruits1.0@gmail.com"
        message = "Dear all, You have been added as the member of Recruitment Committee for the Vacancy {} , Vacancy Id {}".format(vacancy.title,vacancy.id)
        send_mail(
            message_name,#subject
            message,#message
            message_email,#from email
            [first_user.email,second_user.email,third_user.email,fourth_user.email,fifth_user.email]#to email
        )
        return Response({"mssg" : "committee updated successfully"}, status = 204)